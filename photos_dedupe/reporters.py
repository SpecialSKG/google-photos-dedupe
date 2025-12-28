"""
Report generation for deduplication results.
"""

import csv
import json
import logging
from pathlib import Path
from typing import List
from photos_dedupe.dedupe import DuplicateGroup

logger = logging.getLogger(__name__)

try:
    import numpy as np
except Exception:
    np = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

from photos_dedupe.date_utils import get_capture_year, get_date_source_used, infer_account

def _json_default(o):
    if np is not None:
        if isinstance(o, np.integer):
            return int(o)
        if isinstance(o, np.floating):
            return float(o)
        if isinstance(o, np.bool_):
            return bool(o)
    raise TypeError(f"Object of type {o.__class__.__name__} is not JSON serializable")

class Reporter:
    """Generate reports in multiple formats."""
    
    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.reports_dir = self.output_dir / "REPORTS"
        self.reports_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_csv_report(self, groups: List[DuplicateGroup]) -> str:
        """
        Generate CSV report of duplicate groups.
        
        Returns:
            Path to the generated CSV file
        """
        csv_path = self.reports_dir / "dedupe_report.csv"
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'group_id',
                'detection_type',
                'winner_path',
                'winner_sha256',
                'winner_width',
                'winner_height',
                'winner_bytes',
                'duplicate_path',
                'duplicate_sha256',
                'duplicate_width',
                'duplicate_height',
                'duplicate_bytes',
                'phash_distance',
                'reason'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for group in groups:
                for i, dup_path in enumerate(group.duplicates):
                    dup_meta = group.duplicate_metadata[i]
                    
                    row = {
                        'group_id': group.group_id,
                        'detection_type': group.detection_type,
                        'winner_path': group.winner,
                        'winner_sha256': group.winner_metadata.get('sha256', ''),
                        'winner_width': group.winner_metadata.get('width', ''),
                        'winner_height': group.winner_metadata.get('height', ''),
                        'winner_bytes': group.winner_metadata.get('size', ''),
                        'duplicate_path': dup_path,
                        'duplicate_sha256': dup_meta.get('sha256', ''),
                        'duplicate_width': dup_meta.get('width', ''),
                        'duplicate_height': dup_meta.get('height', ''),
                        'duplicate_bytes': dup_meta.get('size', ''),
                        'phash_distance': group.phash_distance if group.phash_distance is not None else '',
                        'reason': group.reason
                    }
                    
                    writer.writerow(row)
        
        logger.info(f"CSV report generated: {csv_path}")
        return str(csv_path)
    
    def generate_json_report(self, groups: List[DuplicateGroup]) -> str:
        """
        Generate JSON report of duplicate groups.
        
        Returns:
            Path to the generated JSON file
        """
        json_path = self.reports_dir / "dedupe_report.json"
        
        report_data = []
        
        for group in groups:
            group_data = {
                'group_id': group.group_id,
                'detection_type': group.detection_type,
                'winner': {
                    'path': group.winner,
                    'sha256': group.winner_metadata.get('sha256', ''),
                    'width': group.winner_metadata.get('width'),
                    'height': group.winner_metadata.get('height'),
                    'bytes': group.winner_metadata.get('size', 0)
                },
                'duplicates': [
                    {
                        'path': dup_path,
                        'sha256': group.duplicate_metadata[i].get('sha256', ''),
                        'width': group.duplicate_metadata[i].get('width'),
                        'height': group.duplicate_metadata[i].get('height'),
                        'bytes': group.duplicate_metadata[i].get('size', 0)
                    }
                    for i, dup_path in enumerate(group.duplicates)
                ],
                'phash_distance': group.phash_distance,
                'reason': group.reason
            }
            
            report_data.append(group_data)
        
        with open(json_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(report_data, jsonfile, indent=2, ensure_ascii=False, default=_json_default)
        
        logger.info(f"JSON report generated: {json_path}")
        return str(json_path)
    
    def generate_xlsx_report(self, 
                            groups: List[DuplicateGroup],
                            inputs: List[str],
                            date_priority: List[str],
                            unknown_year_dir: str,
                            timezone_mode: str,
                            detected_roots: List[str],
                            total_files: int,
                            unique_files: int,
                            mode: str,
                            action: str) -> str:
        """
        Generate Excel report with SUMMARY sheet and year-based sheets.
        
        Returns:
            Path to the generated Excel file
        """
        if Workbook is None:
            logger.warning("openpyxl not available, skipping Excel report generation")
            return ""
        
        xlsx_path = self.reports_dir / "dedupe_report.xlsx"
        wb = Workbook()
        
        # Convert inputs list to tuple for caching
        inputs_tuple = tuple(inputs)
        
        # ===== SUMMARY SHEET =====
        ws_summary = wb.active
        ws_summary.title = "SUMMARY"
        
        # Header style
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="left", vertical="center")
        
        row = 1
        ws_summary[f'A{row}'] = "GOOGLE PHOTOS DEDUPLICATION REPORT"
        ws_summary[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Configuration section
        ws_summary[f'A{row}'] = "CONFIGURATION"
        ws_summary[f'A{row}'].font = header_font
        row += 1
        ws_summary[f'A{row}'] = "Detection Mode:"
        ws_summary[f'B{row}'] = mode
        row += 1
        ws_summary[f'A{row}'] = "Action:"
        ws_summary[f'B{row}'] = action
        row += 1
        ws_summary[f'A{row}'] = "Date Priority:"
        ws_summary[f'B{row}'] = ", ".join(date_priority)
        row += 1
        ws_summary[f'A{row}'] = "Timezone Mode:"
        ws_summary[f'B{row}'] = timezone_mode
        row += 2
        
        # Detected roots section
        ws_summary[f'A{row}'] = "INPUT DIRECTORIES"
        ws_summary[f'A{row}'].font = header_font
        row += 1
        for root in detected_roots:
            ws_summary[f'A{row}'] = root
            row += 1
        row += 1
        
        # Results section
        ws_summary[f'A{row}'] = "RESULTS"
        ws_summary[f'A{row}'].font = header_font
        row += 1
        ws_summary[f'A{row}'] = "Total files scanned:"
        ws_summary[f'B{row}'] = total_files
        row += 1
        ws_summary[f'A{row}'] = "Unique files:"
        ws_summary[f'B{row}'] = unique_files
        row += 1
        ws_summary[f'A{row}'] = "Duplicate groups:"
        ws_summary[f'B{row}'] = len(groups)
        row += 1
        total_duplicates = sum(len(g.duplicates) for g in groups)
        ws_summary[f'A{row}'] = "Total duplicate files:"
        ws_summary[f'B{row}'] = total_duplicates
        
        # Auto-adjust column widths for summary
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 50
        
        # ===== DATA SHEETS BY YEAR =====
        # Collect all rows and organize by year
        rows_by_year = {}
        
        for group in groups:
            winner_path = Path(group.winner)
            
            # Get year from winner (group's year)
            group_year = get_capture_year(
                winner_path,
                date_priority,
                unknown_year_dir,
                timezone_mode
            )
            
            # Get accounts
            winner_account = infer_account(str(winner_path), inputs_tuple)
            
            # Get date source for winner
            winner_date_source = get_date_source_used(winner_path, date_priority, timezone_mode)
            
            for i, dup_path in enumerate(group.duplicates):
                dup_path_obj = Path(dup_path)
                dup_meta = group.duplicate_metadata[i]
                
                # Get duplicate account and year (for auditing)
                dup_account = infer_account(str(dup_path), inputs_tuple)
                dup_year = get_capture_year(
                    dup_path_obj,
                    date_priority,
                    unknown_year_dir,
                    timezone_mode
                )
                dup_date_source = get_date_source_used(dup_path_obj, date_priority, timezone_mode)
                
                row_data = {
                    'group_id': group.group_id,
                    'detection_type': group.detection_type,
                    'phash_distance': group.phash_distance if group.phash_distance is not None else '',
                    'reason': group.reason,
                    'winner_path': group.winner,
                    'winner_account': winner_account,
                    'winner_year': group_year,
                    'winner_date_source': winner_date_source,
                    'winner_sha256': group.winner_metadata.get('sha256', ''),
                    'winner_width': group.winner_metadata.get('width', ''),
                    'winner_height': group.winner_metadata.get('height', ''),
                    'winner_bytes': group.winner_metadata.get('size', ''),
                    'dup_path': dup_path,
                    'dup_account': dup_account,
                    'dup_year': dup_year,
                    'dup_date_source': dup_date_source,
                    'dup_sha256': dup_meta.get('sha256', ''),
                    'dup_width': dup_meta.get('width', ''),
                    'dup_height': dup_meta.get('height', ''),
                    'dup_bytes': dup_meta.get('size', '')
                }
                
                # Add to year-based collection
                if group_year not in rows_by_year:
                    rows_by_year[group_year] = []
                rows_by_year[group_year].append(row_data)
        
        # Create a sheet for each year
        column_headers = [
            'group_id', 'detection_type', 'phash_distance', 'reason',
            'winner_path', 'winner_account', 'winner_year', 'winner_date_source',
            'winner_sha256', 'winner_width', 'winner_height', 'winner_bytes',
            'dup_path', 'dup_account', 'dup_year', 'dup_date_source',
            'dup_sha256', 'dup_width', 'dup_height', 'dup_bytes'
        ]
        
        # Sort years (put _UNKNOWN at the end)
        sorted_years = sorted(
            rows_by_year.keys(),
            key=lambda y: ('z' + y) if y == unknown_year_dir else y
        )
        
        for year in sorted_years:
            rows = rows_by_year[year]
            
            # Create sheet for this year
            ws = wb.create_sheet(title=str(year))
            
            # Write header row
            for col_idx, header in enumerate(column_headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, header in enumerate(column_headers, 1):
                    ws.cell(row=row_idx, column=col_idx).value = row_data.get(header, '')
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Enable auto-filter
            ws.auto_filter.ref = ws.dimensions
            
            # Auto-adjust column widths (with limits)
            for col_idx, header in enumerate(column_headers, 1):
                col_letter = get_column_letter(col_idx)
                if 'path' in header:
                    ws.column_dimensions[col_letter].width = 60
                elif 'sha256' in header:
                    ws.column_dimensions[col_letter].width = 30
                elif 'account' in header:
                    ws.column_dimensions[col_letter].width = 25
                else:
                    ws.column_dimensions[col_letter].width = 15
        
        # Save workbook
        wb.save(xlsx_path)
        logger.info(f"Excel report generated: {xlsx_path}")
        return str(xlsx_path)
    
    def generate_summary(self, 
                        total_files: int,
                        unique_files: int,
                        duplicate_groups: int,
                        total_duplicates: int,
                        detected_roots: List[str],
                        mode: str,
                        action: str) -> str:
        """
        Generate text summary report.
        
        Returns:
            Path to the generated summary file
        """
        summary_path = self.reports_dir / "run_summary.txt"
        
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("GOOGLE PHOTOS DEDUPLICATION SUMMARY\n")
            f.write("=" * 80 + "\n\n")
            
            f.write("CONFIGURATION\n")
            f.write("-" * 80 + "\n")
            f.write(f"Detection Mode: {mode}\n")
            f.write(f"Action: {action}\n")
            f.write("\n")
            
            f.write("DETECTED PHOTO FOLDERS\n")
            f.write("-" * 80 + "\n")
            for i, root in enumerate(detected_roots, 1):
                f.write(f"{i}. {root}\n")
            f.write("\n")
            
            f.write("RESULTS\n")
            f.write("-" * 80 + "\n")
            f.write(f"Total files scanned: {total_files}\n")
            f.write(f"Unique files: {unique_files}\n")
            f.write(f"Duplicate groups found: {duplicate_groups}\n")
            f.write(f"Total duplicate files: {total_duplicates}\n")
            f.write("\n")
            
            space_saved = 0  # Could calculate actual space saved
            f.write(f"Space that can be saved: {space_saved} bytes\n")
            f.write("\n")
            
            f.write("OUTPUT STRUCTURE\n")
            f.write("-" * 80 + "\n")
            f.write(f"UNIQUE/      - {unique_files} unique files\n")
            f.write(f"DUPLICATES/  - {total_duplicates} duplicate files\n")
            f.write(f"REPORTS/     - CSV, JSON, and this summary\n")
            f.write(f"LOGS/        - Detailed execution logs\n")
            f.write("\n")
            
            f.write("=" * 80 + "\n")
        
        logger.info(f"Summary report generated: {summary_path}")
        return str(summary_path)
    
    def generate_all_reports(self,
                            groups: List[DuplicateGroup],
                            total_files: int,
                            unique_files: int,
                            detected_roots: List[str],
                            mode: str,
                            action: str,
                            config=None) -> None:
        """Generate all reports (CSV, JSON, Excel, and summary).
        
        Args:
            groups: List of DuplicateGroup objects
            total_files: Total number of files scanned
            unique_files: Number of unique files
            detected_roots: List of detected root directories
            mode: Detection mode used
            action: Action performed
            config: Configuration object (needed for Excel and conditional generation)
        """
        # Conditionally generate reports based on config flags
        if config:
            if config.reports_csv:
                self.generate_csv_report(groups)
            
            if config.reports_json:
                self.generate_json_report(groups)
            
            if config.reports_xlsx:
                self.generate_xlsx_report(
                    groups=groups,
                    inputs=config.inputs,
                    date_priority=config.date_source_priority,
                    unknown_year_dir=config.unknown_year_dir,
                    timezone_mode=config.timezone_mode,
                    detected_roots=detected_roots,
                    total_files=total_files,
                    unique_files=unique_files,
                    mode=mode,
                    action=action
                )
        else:
            # Backward compatibility: generate all reports if no config
            self.generate_csv_report(groups)
            self.generate_json_report(groups)
        
        # Always generate summary
        duplicate_count = sum(len(g.duplicates) for g in groups)
        self.generate_summary(
            total_files=total_files,
            unique_files=unique_files,
            duplicate_groups=len(groups),
            total_duplicates=duplicate_count,
            detected_roots=detected_roots,
            mode=mode,
            action=action
        )

